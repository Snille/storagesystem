from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
INVENTORY_PATH = ROOT / "data" / "inventory.json"

REQUIRED_HEADERS = {
    "Namn",
    "Plats",
    "Box-ID",
    "Plats-ID",
}


def load_inventory() -> dict[str, Any]:
    return json.loads(INVENTORY_PATH.read_text(encoding="utf-8"))


def save_inventory(data: dict[str, Any]) -> None:
    INVENTORY_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_keywords(value: str) -> list[str]:
    seen: list[str] = []
    for part in value.split(","):
        keyword = part.strip()
        if keyword and keyword not in seen:
            seen.append(keyword)
    return seen


def parse_location_display(raw_location: str) -> str:
    trimmed = raw_location.strip()

    ivar_match = re.match(r"^Ivar:\s*([A-Z0-9-]+)\s*,\s*Hylla:\s*(\d+)\s*,\s*Plats:\s*(\d+)([A-Z])?$", trimmed, re.IGNORECASE)
    if ivar_match:
        unit = ivar_match.group(1).upper()
        shelf = ivar_match.group(2)
        slot = ivar_match.group(3)
        variant = ivar_match.group(4).upper() if ivar_match.group(4) else ""
        return f"IVAR:{unit}:H{shelf}:P{slot}{f':{variant}' if variant else ''}"

    bench_match = re.match(
        r"^Bänk:\s*([A-Z0-9-]+)\s*,\s*Yta:\s*(Ovanpå|Under)\s*,\s*Plats:\s*(\d+)([A-Z])?$",
        trimmed,
        re.IGNORECASE,
    )
    if bench_match:
        unit = bench_match.group(1).upper()
        row = "UNDER" if bench_match.group(2).lower() == "under" else "TOP"
        slot = bench_match.group(3)
        variant = bench_match.group(4).upper() if bench_match.group(4) else ""
        return f"BENCH:{unit}:{row}:P{slot}{f':{variant}' if variant else ''}"

    cabinet_match = re.match(r"^Skåp:\s*([A-Z0-9-]+)\s*,\s*Hylla:\s*(\d+)\s*,\s*Plats:\s*(\d+)([A-Z])?$", trimmed, re.IGNORECASE)
    if cabinet_match:
        unit = cabinet_match.group(1).upper()
        shelf = cabinet_match.group(2)
        slot = cabinet_match.group(3)
        variant = cabinet_match.group(4).upper() if cabinet_match.group(4) else ""
        return f"CABINET:{unit}:H{shelf}:P{slot}{f':{variant}' if variant else ''}"

    legacy_match = re.match(r"^([A-Z])-H(\d+)-P(\d+)(?:-([A-Z]))?$", trimmed, re.IGNORECASE)
    if legacy_match:
        unit = legacy_match.group(1).upper()
        shelf = legacy_match.group(2)
        slot = legacy_match.group(3)
        variant = legacy_match.group(4).upper() if legacy_match.group(4) else ""
        return f"IVAR:{unit}:H{shelf}:P{slot}{f':{variant}' if variant else ''}"

    return trimmed


def load_catalog_rows(xlsx_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active
    headers = [normalize_text(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    missing_headers = sorted(REQUIRED_HEADERS - set(headers))
    if missing_headers:
        raise RuntimeError(
            "Excel-filen använder inte exportformatet från appen. Saknade kolumner: "
            + ", ".join(missing_headers)
        )

    header_index = {header: index for index, header in enumerate(headers)}
    rows: list[dict[str, str]] = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        name = normalize_text(values[header_index["Namn"]])
        box_id = normalize_text(values[header_index["Box-ID"]])
        location_id = normalize_text(values[header_index["Plats-ID"]])
        raw_location = normalize_text(values[header_index["Plats"]])

        if not name and not box_id and not location_id:
            continue

        normalized_location_id = location_id or parse_location_display(raw_location)
        if not box_id:
            raise RuntimeError("Box-ID saknas i en eller flera rader. Import kräver exportformatet från appen.")
        if not normalized_location_id:
            raise RuntimeError(f"Plats-ID saknas eller kunde inte tolkas för raden med box {box_id}.")

        rows.append(
            {
                "label": name or box_id,
                "box_id": box_id,
                "location_id": normalized_location_id,
                "box_notes": normalize_text(values[header_index.get("Beskrivning", -1)]) if "Beskrivning" in header_index else "",
                "summary": normalize_text(values[header_index.get("Sammanfattning", -1)]) if "Sammanfattning" in header_index else "",
                "keywords": normalize_text(values[header_index.get("Nyckelord", -1)]) if "Nyckelord" in header_index else "",
                "session_notes": normalize_text(values[header_index.get("Sessionsanteckningar", -1)]) if "Sessionsanteckningar" in header_index else "",
                "session_id": normalize_text(values[header_index.get("Session-ID", -1)]) if "Session-ID" in header_index else "",
                "created_at": normalize_text(values[header_index.get("Skapad", -1)]) if "Skapad" in header_index else "",
                "updated_at": normalize_text(values[header_index.get("Uppdaterad", -1)]) if "Uppdaterad" in header_index else "",
            }
        )

    return rows


def merge_catalog(xlsx_path: Path) -> dict[str, int]:
    rows = load_catalog_rows(xlsx_path)
    inventory = load_inventory()
    timestamp = now_iso()

    boxes: list[dict[str, Any]] = inventory["boxes"]
    sessions: list[dict[str, Any]] = inventory["sessions"]
    photos: list[dict[str, Any]] = inventory["photos"]

    boxes_by_id = {box["boxId"]: box for box in boxes}
    sessions_by_id = {session["sessionId"]: session for session in sessions}

    created_boxes = 0
    updated_boxes = 0
    created_sessions = 0
    updated_sessions = 0

    for row in rows:
        existing_box = boxes_by_id.get(row["box_id"])
        created_at = row["created_at"] or timestamp
        updated_at = row["updated_at"] or timestamp

        if existing_box:
            existing_box["label"] = row["label"]
            existing_box["currentLocationId"] = row["location_id"]
            existing_box["notes"] = row["box_notes"] or None
            existing_box["updatedAt"] = updated_at
            box = existing_box
            updated_boxes += 1
        else:
            box = {
                "boxId": row["box_id"],
                "label": row["label"],
                "currentLocationId": row["location_id"],
                "notes": row["box_notes"] or None,
                "createdAt": created_at,
                "updatedAt": updated_at,
            }
            boxes.append(box)
            boxes_by_id[box["boxId"]] = box
            created_boxes += 1

        for session in sessions:
            if session["boxId"] == box["boxId"]:
                session["isCurrent"] = False

        session_id = row["session_id"] or f"IMPORT-{box['boxId']}"
        summary = row["summary"] or row["box_notes"] or f"Importerad från katalog: {row['label']}"
        session_notes = row["session_notes"] or "Importerad från Excel-export."
        item_keywords = parse_keywords(row["keywords"])

        existing_session = sessions_by_id.get(session_id)
        if existing_session:
            existing_session["boxId"] = box["boxId"]
            existing_session["createdAt"] = existing_session.get("createdAt") or created_at
            existing_session["summary"] = summary
            existing_session["itemKeywords"] = item_keywords
            existing_session["notes"] = session_notes
            existing_session["isCurrent"] = True
            updated_sessions += 1
        else:
            session = {
                "sessionId": session_id,
                "boxId": box["boxId"],
                "createdAt": created_at,
                "summary": summary,
                "itemKeywords": item_keywords,
                "notes": session_notes,
                "isCurrent": True,
            }
            sessions.append(session)
            sessions_by_id[session_id] = session
            created_sessions += 1

    inventory["boxes"] = sorted(boxes, key=lambda item: item["boxId"])
    inventory["sessions"] = sorted(sessions, key=lambda item: item["sessionId"])
    inventory["photos"] = photos
    save_inventory(inventory)

    return {
        "catalog_rows": len(rows),
        "created_boxes": created_boxes,
        "updated_boxes": updated_boxes,
        "created_sessions": created_sessions,
        "updated_sessions": updated_sessions,
        "total_boxes": len(inventory["boxes"]),
        "total_sessions": len(inventory["sessions"]),
        "total_photos": len(inventory["photos"]),
    }


def main() -> int:
    if len(sys.argv) < 2:
        raise SystemExit("Ange sökvägen till en exporterad Excel-fil som första argument.")

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.is_absolute():
        xlsx_path = (ROOT / xlsx_path).resolve()
    if not xlsx_path.exists():
        raise SystemExit(f"Excel-filen hittades inte: {xlsx_path}")

    summary = merge_catalog(xlsx_path)
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
